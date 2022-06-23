"""Fill the excel file with delta of GSM parameters."""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def get_diff_values(cell_name, diff_values, cells_by_band):
    """
    Get parameter value and baseline value accoding to GSM cell frequency band.

    Args:
        cell_name: string
        diff_values: tuple
        cells_by_band: dict

    Returns:
        tuple
    """
    if isinstance(diff_values[0], tuple):
        cell_band = cells_by_band[cell_name]
        if cell_band == 'GSM900':
            parameter_values = (diff_values[0][0], diff_values[1][0])
        elif cell_band == 'GSM1800':
            parameter_values = (diff_values[0][1], diff_values[1][1])
    else:
        parameter_values = diff_values
    return parameter_values


def fill_inconsistencies(sheet, bsc, parameters_diff, cells_by_band):
    """
    Fill the parameter inconsistencies to a given excel sheet.

    Args:
        sheet: Workbook sheet
        bsc: string
        parameters_diff: dict
        cells_by_band: dict
    """
    headers = ['BSC', 'Cell', 'Parameter', 'Current Value', 'Baseline Value']
    for header in headers:
        column_index = headers.index(header) + 1
        sheet.cell(row=1, column=column_index, value=header)
        sheet.column_dimensions[get_column_letter(column_index)].width = 20

    params_row = 2

    for cell_name, params_diff in parameters_diff.items():
        for parameter_name, cell_parameters_diff in params_diff.items():
            parameter_values = get_diff_values(cell_name, cell_parameters_diff, cells_by_band)

            sheet.cell(row=params_row, column=headers.index('BSC') + 1, value=bsc)
            sheet.cell(row=params_row, column=headers.index('Cell') + 1, value=cell_name)
            sheet.cell(
                row=params_row,
                column=headers.index('Parameter') + 1,
                value=parameter_name,
            )
            sheet.cell(
                row=params_row,
                column=headers.index('Current Value') + 1,
                value=parameter_values[0],
            )
            sheet.cell(
                row=params_row,
                column=headers.index('Baseline Value') + 1,
                value=parameter_values[1],
            )
            params_row += 1


def fill_baseline(sheet, baseline):
    """
    Fill baseline values to a given sheet.

    Args:
        sheet: Workbook sheet
        baseline: dict
    """
    baseline_headers = [
        'Parameter',
        'Common Baseline Value',
        'GSM900 Baseline Value',
        'GSM1800 Baseline Value',
    ]
    for header in baseline_headers:
        column_index = baseline_headers.index(header) + 1
        sheet.cell(row=1, column=column_index, value=header)
        sheet.column_dimensions[get_column_letter(column_index)].width = 20

    params_row = 2
    for baseline_parameter, baseline_value in baseline.items():
        sheet.cell(
            row=params_row,
            column=baseline_headers.index('Parameter') + 1,
            value=baseline_parameter,
        )
        if isinstance(baseline_value, tuple):
            sheet.cell(
                row=params_row,
                column=baseline_headers.index('GSM900 Baseline Value') + 1,
                value=baseline_value[0],
            )
            sheet.cell(
                row=params_row,
                column=baseline_headers.index('GSM1800 Baseline Value') + 1,
                value=baseline_value[1],
            )
        else:
            sheet.cell(
                row=params_row,
                column=baseline_headers.index('Common Baseline Value') + 1,
                value=baseline_value,
            )
        params_row += 1


def fill_gsm_report(bsc, parameters_diff, baseline, cells_by_band):
    """
    Fill excel file with delta of GSM parameters.

    Args:
        bsc: string
        parameters_diff: dict
        baseline: dict
        cells_by_band: dict
    """
    logs_path = 'baseline_checker/logs'
    work_book = Workbook()
    delta_sheet = work_book.active

    delta_sheet.title = 'Inconsistencies'
    fill_inconsistencies(delta_sheet, bsc, parameters_diff, cells_by_band)

    baseline_sheet = work_book.create_sheet('Baseline')
    fill_baseline(baseline_sheet, baseline)

    work_book.save(f'{logs_path}/{bsc}-inconsistencies.xlsx')
